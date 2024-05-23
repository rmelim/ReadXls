"""
Nombre de la clase: ReadXls.
Creada originalmente por: Rui Duarte dos Santos Melim

Dependencias:
    Módulo openpyxl.py -- https://pypi.org/project/openpyxl/

Versiones requeridas:
    Python 3.6 o superior
    openpyxl 3.1.2 o superior

La clase ReadXls, a través de las funcionalidades del módulo openpyxl, 
sirve para leer los datos de una hoja de un archivo específico de Excel.
"""

from openpyxl import load_workbook


class ReadXls:
    """
    Permite leer una hoja activa de un archivo de Excel específico.

    Propiedades de la clase:
        * workbook - Contiene el objeto instanciado de openpyxl indicado
        en el argumento filename.

        * worksheets - Contiene unalista de las hojas en workbook.

        * sheet - Contiene la hoja de cálculo activa, por defecto la primera
        en la lista.

        * coords - Contiene una lista con el rango de las celdas a leer en el
        formato ["primera celda", "última celda"], ejemplo: ["C7", "R23"].
        Por defecto, las coordenadas de la primera celda de sheet.

        * cells - Contiene una lista de listas con los valores del grupo de
        celdas según el rango especificado en la propiedad coords. El orden
        se establece en filas/columnas. Esta propiedad es de sólo lectura.
    """

    def __init__(self, filename=""):
        if filename != "":
            self.workbook = load_workbook(filename, data_only=True)
            self.worksheets = self.workbook.sheetnames
            self._sheet = self.workbook[self.worksheets[0]].title
            self._coords = ["A1", "A1"]
            self._cells = ReadXls._read_cells(self.workbook[self.sheet], self.coords)

    @property
    def sheet(self):
        """Getter, y posterior Setter, para propiedad sheet."""
        return self._sheet

    @sheet.setter
    def sheet(self, value):
        self._sheet = self.workbook[value].title

    @property
    def coords(self):
        """Getter, y posterior Setter, para propiedad coords."""
        return self._coords

    @coords.setter
    def coords(self, value):
        self._coords = value
        self._cells = ReadXls._read_cells(self.workbook[self.sheet], self.coords)

    @property
    def cells(self):
        """Getter para propiedad cells."""
        return self._cells

    @staticmethod
    def _read_cells(sheet, coords):
        """
        Método interno a la clase que lee el grupo de celdas indicado
        en coords y devuelve una lista de valores.

        Arguments:
            sheet -- Hoja activa a extraer los datos. Tipo: string.
            coords -- Coordenadas del rango de celdas a extraer.
            Tipo: lista de string.

        Returns:
            Lista de listas (matriz) con los valores de las celdas leídas.
        """
        cells = []
        for row in sheet[coords[0] : coords[1]]:
            cell = [i.value for i in row]
            cells.append(cell)
        return cells


# ***************************
# Prueba del código.
# ***************************
if __name__ == "__main__":

    PATH = __file__.replace("readxls.py", "")

    my_excel = ReadXls(PATH + "Test.xlsx")

    print("Lista de hojas\n", my_excel.worksheets, "\n")

    print("Valores por defecto:")
    print(my_excel.coords)
    print(my_excel.sheet)
    print(my_excel.cells)

    print("\nValores de propiedad cambiados:")
    my_excel.sheet = "Analisis de Ventas"
    my_excel.coords = ["A2", "E6"]
    print(my_excel.coords)
    print(my_excel.sheet)
    print(my_excel.cells)
